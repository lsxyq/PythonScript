#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# Author:Leslie-x
import requests, re, random, sys
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook

import pymysql


# 连接数据库
def get_connect():
    conn = pymysql.connect(
        host='127.0.0.1',
        port=3306,
        user='root',
        passwd='root',
        db='seotest',
        charset="utf8")
    # 获取游标对象
    cursor = conn.cursor()
    return conn, cursor


# 解析excel文件，取出所有的url
def get_urls(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    urls = []
    for cell in list(sheet.columns)[1]:
        if cell != sheet['B1']:
            urls.append(cell.value)
    return wb, urls


# 伪造请求，取得html页面
def get_html(url):
    # 定义http的请求Header
    headers = {}
    # random.randint(1,99) 为了生成1到99之间的随机数，让UserAgent变的不同。
    headers[
        'User-Agent'] = "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537." + str(
        random.randint(1, 99))
    # Referer地址使用待查询的网址
    headers['Referer'] = "http://seo.chinaz.com/" + url + "/"
    html = ''
    try:
        html = requests.get("http://seo.chinaz.com/" + url + "/", headers=headers, timeout=5).text
    except Exception:
        pass
    # 分析页面元素，提取需要字段
    return html


# 利用BeautifulSoup模块从html页面中提取数据
def get_data(html, url):
    if not html:
        return url, 0
    soup = bs(html, "lxml")
    p_tag = soup.select("p.ReLImgCenter")[0]
    src = p_tag.img.attrs["src"]
    regexp = re.compile(r'^http:.*?(\d).gif')
    br = regexp.findall(src)[0]
    return url, br


# 主程序
if __name__ == "__main__":
    #命令行执行脚本文件，获取excel文件路径
    file_path = sys.argv[1]
    #获取URL列表和excle工作簿
    wb, urls = get_urls(file_path)
    #获取数据库连接和游标
    conn, cursor = get_connect()
    #获取工作簿当前工作sheet
    sheet = wb.active
    #数据库插入语句
    sql_insert = '''insert into website_weight(main_url, website_weight) values (%s, %s)'''

    for row, url in enumerate(urls):
        if not url: continue
        html = get_html(url)
        data = get_data(html, url)

        # 插入数据到数据库
        cursor.execute(sql_insert, data)
        # 插入数据到Excel表中
        cell = sheet.cell(row=row + 2, column=3)
        cell.value = data[1]
        # 终端打印插入的数据
        print(data)
    conn.commit()
    conn.close()
    wb.save(file_path)
    wb.close()

# cmd命令：python3 F:\算法与结构\website_weight.py F:\website.xlsx
